#! /usr/bin/env python3
#
# Command line python script to syntax check SPIF formatted patent data
#
# Usage: spiffy.py inputfile.xslx
#
# Specification verion: 0.2.1
# Specification is at: https://spif.group
#
# (C) 2020, 2021 Erik Oliver
#
# Permission is hereby granted, free of charge, to any person obtaining
# a copy of this software and associated documentation files (the
# "Software"), to deal in the Software without restriction, including
# without limitation the rights to use, copy, modify, merge, publish,
# distribute, sublicense, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to
# the following conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
# MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY
# CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
# TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
# SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.


from openpyxl import Workbook, load_workbook #https://openpyxl.readthedocs.io/en/stable/tutorial.html
import sys
import re

# Key terms defined in the specification
SHEETNAME = 'Master Data - SPIF'
APPNUM = 'Application Number - SPIF'
PUBNUM = 'Publication Number - SPIF'
SUPPORTEDAPPCOUNTRIES = ['US', 'KR', 'JP', 'CN', 'EP', 'WO']
SUPPORTEDPUBCOUNTRIES = ['US', 'KR', 'JP', 'CN', 'EP', 'WO']

# NOTES: Since only 2000-forward is supported for US pubs and WO pubs/apps looks for 20## instead any 4-digit year

def checkappnum(worksheet, row, column, errorcolumn):
    appnum = worksheet.cell(row=row,column=column).value
    country = appnum[:2]
    # skip over countries we do not support
    if not(country in SUPPORTEDAPPCOUNTRIES):
        worksheet.cell(row=row, column=errorcolumn, value='Unsupported country {}'.format(country))
        return

    if (country == 'US'):
        if not (re.match("^US\d{8}$",appnum)):
            worksheet.cell(row=row, column=errorcolumn, value='US Application Numbers should be US######## (US followed by 8-digits)')
        else:
            worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'EP'):
        if not (re.match("^EP\d{8}$",appnum)):
            worksheet.cell(row=row, column=errorcolumn, value='EP Application Numbers should be EP######## (EP followed by 8-digits)')
        else:
            worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'JP'):
        m = re.match("^JP(\d{4})\d{6}$",appnum)
        if not m:
            worksheet.cell(row=row, column=errorcolumn, value='JP Application Numbers should be JPYYYY###### (6-digits)')
        else:
            if(int(m.group(1)) < 2000):
                worksheet.cell(row=row, column=errorcolumn, value='Year predates 2000 not checked')
            else:
                worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'WO'):
        m = re.match("^(WO(\d{4})[A-Z]{2}\d{6})$",appnum)
        if not m:
            worksheet.cell(row=row, column=errorcolumn, value='WO Application Numbers should be WOYYYYCC###### (6-digits)')
        else:
            if(int(m.group(2)) < 2000):
                worksheet.cell(row=row, column=errorcolumn, value='Year predates 2000 not checked')
            else:
                worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'CN'):
        m = re.match("^CN(\d{4})[1289]\d{6,7}$",appnum)
        if not m:
            # note no checks here for Aug 2007 cutoff since date is not known TODO - flag 6/7 digits for non 2007 years
            worksheet.cell(row=row, column=errorcolumn, value='CN Application Numbers should be CNYYYY followed by 1, 2, 8, or 9, and then 6 digits pre Aug 2007 and 7 digits post Aug 2007')
        else:
            if(int(m.group(1)) < 2000):
                worksheet.cell(row=row, column=errorcolumn, value='Year predates 2000 not checked')
            else:
                worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'KR'):
        m = re.match("^KR[12]0(\d{4})\d{7}$",appnum)
        if not m:
            worksheet.cell(row=row, column=errorcolumn, value='KR Application Numbers should be KR10YYYY####### or KR20YYYY####### (7-digits in both)')
        else:
            if(int(m.group(1)) < 2000):
                worksheet.cell(row=row, column=errorcolumn, value='Year predates 2000 not checked')
            else:
                worksheet.cell(row=row, column=errorcolumn, value='OK')
    else:
        worksheet.cell(row=row, column=errorcolumn, value='Not yet implemented')


def checkpubnum(worksheet, row, column, errorcolumn):
    pubnum = worksheet.cell(row=row,column=column).value
    country = pubnum[:2]
    # skip over countries we do not support
    if not(country in SUPPORTEDPUBCOUNTRIES):
        worksheet.cell(row=row, column=errorcolumn, value='Unsupported country {}'.format(country))
        return

    if (country == 'US'):
        # try to match US allowed styles for pub numbers
        # https://www.uspto.gov/learning-and-resources/support-centers/electronic-business-center/kind-codes-included-uspto-patent
        # note because this regex requires 2000 in the years no additional 2000 checkings
        if not(re.match("^(US20\d{2}\d{7}A\d)|(US[0-1]?\d{7}[A-BCEFJKO][1-9]?|USRE\d{5}E\d?)$",pubnum)):
            worksheet.cell(row=row, column=errorcolumn, value='US numbers should be USYYYY#######KK (US followed by 4-digit year, 7-digit pub, kind code) or US#######KK/US########KK (7 or 8 digit pub) or USRE#####E or USRE#####E#)')
        else:
            worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'EP'):
        if not (re.match("^EP\d{7}[A-B][1-9]?$",pubnum)):
            worksheet.cell(row=row, column=errorcolumn, value='EP numbers should be EP#######KK (EP followed by 7-digits, followed by kind code)')
        else:
            worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'WO'):
        # note because this regex requires 2000 in the years no additional 2000 checkings
        if not (re.match("^WO20\d{2}\d{6}[A][1-9]$",pubnum)):
            worksheet.cell(row=row, column=errorcolumn, value='WO numbers should be WOYYYY#######KK (WO followed by 4-digit year, by 6-digits, followed by kind code)')
        else:
            worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'CN'):
        if not (re.match("^(CN[12]\d{6})|(CN[12]\d{8}[A-Z]\d$)",pubnum)):
            # note the Aug 2007 cut off for 6 vs. 8 digits is not tested TODO - add warning/test if not 2007
            worksheet.cell(row=row, column=errorcolumn, value='CN numbers should be CN followed by 1 or 2, and either 6 or 8 digits then kind code')
        else:
            worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'JP'):
        m = re.match("^(?:JP(\d{4})\d{6}[A-Z]\d)|(?:JP\d{6}[A-Z]\d)|(?:JP(\d{4})\d{6}U)|(?:JP\d{6}U)$",pubnum)
        if not m:
            # note the Aug 2007 cut off for 6 vs. 8 digits is not tested TODO - add warning/test if not 2007
            worksheet.cell(row=row, column=errorcolumn, value='JP numbers should be JPYYYY######KK or JP######KK or JPYYYY#####U or JP#####U (all 6-digits)')
        else:
            year = -1
            if(m.group(1)):
                year = int(m.group(1))
            elif (m.group(2)):
                year = int(m.group(2))
            if (year != -1 and year < 2000):
                worksheet.cell(row=row, column=errorcolumn, value='Year predates 2000 not checked')
            else:
                worksheet.cell(row=row, column=errorcolumn, value='OK')
    elif (country == 'KR'):
        m = re.match("^(?:KR(\d{4})\d{7}[AU])|(?:KR[12]0(\d{4})\d{7}[AU])|(?:KR[12]0\d{7}[BY]\d)$",pubnum)
        if not m:
            # note the Aug 2007 cut off for 6 vs. 8 digits is not tested TODO - add warning/test if not 2007
            worksheet.cell(row=row, column=errorcolumn, value='KR numbers pre-2004 apps should be KRYYYY#######K (7 digits and A or U kind); post-2004 KR10YYYY#######K or KR20YYYY#######K (7-digits and A or U kind), or KR10########B# or KR20#######Y#')
        else:
            year = -1
            if(m.group(1)):
                year = int(m.group(1))
            elif (m.group(2)):
                year = int(m.group(2))
            if (year != -1 and year < 2000):
                worksheet.cell(row=row, column=errorcolumn, value='Year predates 2000 not checked')
            else:
                worksheet.cell(row=row, column=errorcolumn, value='OK')
    else:
        worksheet.cell(row=row, column=errorcolumn, value='Not yet implemented')

def processinputfile(inputfile):
    # open the the Excel Workbook and find the 'Master Data' sheet and then the 2 key columns
    positions = {APPNUM: -1,  PUBNUM: -1}

    try:
        wb = load_workbook(inputfile)
    except:
        print("Error: openpyxl library was unable to load '{}'\nDouble check that your file can open in Excel".format(inputfile))
        exit(-1)

    # Look at the sheetnames for 'Master Data'
    try:
        mysheet = wb[SHEETNAME]
    except:
        print("Could not find a sheet named '{}' in the workbook.".format(SHEETNAME))
        exit(-1)

    # Look at Row 1 for the column names
    for position, cell in enumerate(mysheet[1], start = 1):
        if (cell.value == APPNUM):
            positions[APPNUM] = position
        elif (cell.value == PUBNUM):
            positions[PUBNUM] = position

    if(positions[APPNUM] == -1 or positions[PUBNUM] == -1):
        if(positions[APPNUM] == -1):
            print("Error: Could not find a column named '{}' in the first row of the sheet '{}'".format(APPNUM,SHEETNAME))
        if(positions[PUBNUM] == -1):
            print("Error: Could not find a column named '{}' in the first row of the sheet '{}'".format(PUBNUM,SHEETNAME))
        exit(-1)
    print("Basic checking complete: sheet name is correct and the two required column names are correct, now checking the contents")

    # add two columns to the right of the existing columns to add results of the analysis
    lastpos = len(mysheet[1]) # we will write in the last column onward
    mysheet.cell(row=1, column=lastpos+1, value='Application Number Errors')
    mysheet.cell(row=1, column=lastpos+2, value='Publication Number Errors')

    # since we only want to look at a handful of cells we will use a more traditional for loop
    numrows = len(mysheet['A']) # get the size of the first column
    if (numrows <= 1):
        print("Error: There does not appear to be data in column A other than the headings, only {} row(s) founds in total".format(numrows))
        exit(-1)

    # Review the APPNUM and PUBNUM in each of the data rows
    for i in range(2, numrows+1):
        checkappnum(mysheet,i,positions[APPNUM], lastpos+1)
        checkpubnum(mysheet,i,positions[PUBNUM], lastpos+2)

    wb.save("results.xlsx") # TODO: Implement better save file name like "inputfile-results.xlsx"
    print("Contents check, results file written.")

def main(argv):
    # check there is at least one arguement and that is appears to be an XSLX file
    inputfile = ''

    if len(sys.argv) != 2:
        print("Usage: ", sys.argv[0], " inputfile.xlsx" )
        exit(-1)

    inputfile = sys.argv[1]

    if not(re.match(r".*(\.xlsx)$",inputfile)):
        print("Expected an '.xlsx' file got '{}'".format(inputfile) )
        exit(-1)

    # process the inputfile
    processinputfile(inputfile)

if __name__ == "__main__":
   main(sys.argv[1:])
